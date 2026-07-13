#!/usr/bin/env python3
"""
extract_json_to_xlsx.py — convert a §8 EXTRACT object into a workbook,
one sheet per top-level key.

Usage:
    python3 extract_json_to_xlsx.py INPUT.json [OUTPUT.xlsx]

If OUTPUT is omitted, it writes alongside INPUT with a .xlsx extension.

Design choices (kept faithful to the §8 contract):
- Each top-level key -> one sheet, structure preserved 1:1.
- List-of-dicts -> record table; header is the union of keys in first-seen
  order (rows may be ragged across the array).
- Dict          -> vertical key/value sheet (metadata blocks like run_meta,
  gold_scoring, locate_reconciliation are not observations).
- Bare scalar   -> single-cell sheet (e.g. schema_stress_note).
- Any cell value that is itself a list/dict is JSON-stringified so it
  round-trips losslessly instead of being coerced or dropped by the writer.
  This matters for cross-reference arrays such as
  coverage_report.evidence_rows and confidence_reasons.evidence_rows.
"""
import json
import sys
from pathlib import Path
from openpyxl import Workbook
from openpyxl.utils import get_column_letter


def cellify(v):
    # Stringify containers; leave scalars (incl. None) native so Excel types them.
    if isinstance(v, (list, dict)):
        return json.dumps(v, ensure_ascii=False)
    return v


def convert(src: Path, out: Path) -> list[str]:
    with open(src, encoding="utf-8") as f:
        data = json.load(f)

    if not isinstance(data, dict):
        raise ValueError(f"Expected a top-level object, got {type(data).__name__}")

    wb = Workbook()
    wb.remove(wb.active)

    for key, val in data.items():
        ws = wb.create_sheet(key[:31])  # Excel caps sheet names at 31 chars.

        if isinstance(val, list) and val and isinstance(val[0], dict):
            header = []
            for row in val:
                for c in row:
                    if c not in header:
                        header.append(c)
            ws.append(header)
            for row in val:
                ws.append([cellify(row.get(c)) for c in header])

        elif isinstance(val, list):
            ws.append([key])
            for item in val:
                ws.append([cellify(item)])

        elif isinstance(val, dict):
            ws.append(["key", "value"])
            for k, v in val.items():
                ws.append([k, cellify(v)])

        else:  # bare scalar
            ws.append([key])
            ws.append([cellify(val)])

        # Light autofit, capped for readability.
        for col in ws.columns:
            width = max((len(str(c.value)) for c in col if c.value is not None), default=10)
            ws.column_dimensions[get_column_letter(col[0].column)].width = min(max(width + 2, 10), 80)
        ws.freeze_panes = "A2"

    wb.save(out)
    return wb.sheetnames


if __name__ == "__main__":
    if len(sys.argv) < 2:
        sys.exit("Usage: python3 extract_json_to_xlsx.py INPUT.json [OUTPUT.xlsx]")
    src = Path(sys.argv[1])
    out = Path(sys.argv[2]) if len(sys.argv) > 2 else src.with_suffix(".xlsx")
    sheets = convert(src, out)
    print(f"Wrote {out}")
    print("Sheets:", sheets)
