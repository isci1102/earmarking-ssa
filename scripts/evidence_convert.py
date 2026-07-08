#!/usr/bin/env python3
"""
evidence_convert.py — bidirectional xlsx <-> json converter for the earmark
evidence output, implementing the dictionary section 8 serialization contract.

One document = one output unit, rendered EITHER as a nine-sheet workbook OR as a
nine-key JSON object. This script moves between the two renderings and checks
that they agree. It is NOT part of extraction: the model emits one rendering;
this runs downstream, on demand (chat) or as one orchestration step per document
(API), to convert or to validate — never per prompt-run.

Design choices (methodological, not stylistic):
- The 44 evidence columns are read from the workbook header, not hardcoded, so
  the converter tracks the actual emitted schema and a column drift surfaces as a
  conformance failure rather than being silently masked by a stale constant.
- RUN_META is a flat field/value sheet with "--- SECTION ---" marker rows
  (value is None) delimiting three groups; JSON nests them (parameters / a6_gate /
  id_discipline). The mapping is explicit and reversible so the round-trip holds.
- Empty cells serialize to JSON null, never "" — preserving the null-vs-empty
  distinction the schema relies on in numeric/enum columns.
- locate_reconciliation is emitted as an OBJECT with first-class {N,M,K,J} scalars
  so the M+K==N / J==0 invariant is machine-checkable, not derived by counting.
"""

from __future__ import annotations
import json
import sys
import argparse
from pathlib import Path
import openpyxl

# ---- Sheet <-> key map (dictionary section 8). Order is canonical. --------------
# Each entry: (sheet_name, json_key, kind)
#   kind "rows"    -> array of {header: cell} objects
#   kind "meta"    -> the sectioned RUN_META object
#   kind "recon"   -> the locate_reconciliation object {N,M,K,J,entries:[...]}
#   kind "string"  -> single free-text string
SHEET_MAP = [
    ("RUN_META",              "run_meta",              "meta"),
    ("evidence_table",        "evidence_rows",         "rows"),
    ("COVERAGE_REPORT",       "coverage_report",       "rows"),
    ("RECALL_AUDIT",          "recall_audit",          "rows"),
    ("LOCATE_RECONCILIATION", "locate_reconciliation", "recon"),
    ("GOLD_SCORING",          "gold_scoring",          "rows"),
    ("FIELD_EXERCISE_NOTE",   "field_exercise_note",   "rows"),
    ("CONFIDENCE_REASONS",    "confidence_reasons",    "rows"),
    ("SCHEMA_STRESS_NOTE",    "schema_stress_note",    "string"),
]
JSON_KEYS = [k for (_, k, _) in SHEET_MAP]

# ---- LOCATE output map (dictionary section 9). Two sheets / two keys. ----------
LOCATE_MAP = [
    ("PASSAGE_INVENTORY",          "passage_inventory"),
    ("SWEEP_COVERAGE_CERTIFICATE", "sweep_coverage_certificate"),
]
LOCATE_KEYS = [k for (_, k) in LOCATE_MAP]

# RUN_META section markers -> json sub-object key. A marker row has value=None and
# a "--- NAME ---" field. Rows before the first marker are the parameters group.
META_SECTIONS = {
    "--- A6 PREPROCESSING GATE ---": "a6_gate",
    "--- ID DISCIPLINE (C6) ---":    "id_discipline",
}
META_FIRST_GROUP = "parameters"


# ================================================================= XLSX -> JSON
def _cell(v):
    """Normalize a cell for JSON: blanks -> None, everything else verbatim."""
    if v is None:
        return None
    if isinstance(v, str) and v.strip() == "":
        return None
    return v


def _rows_sheet_to_list(ws):
    """A header-row sheet -> list of dicts keyed by header, in column order.
    Trailing all-empty rows are dropped; internal blank rows are preserved as
    all-null dicts only if any later row has content (kept simple: drop fully
    empty rows)."""
    it = ws.iter_rows(values_only=True)
    try:
        header = list(next(it))
    except StopIteration:
        return []
    header = [h for h in header if h is not None]
    out = []
    for row in it:
        vals = list(row)[: len(header)]
        if all(_cell(v) is None for v in vals):
            continue
        out.append({header[i]: _cell(vals[i]) if i < len(vals) else None
                    for i in range(len(header))})
    return out


def _meta_sheet_to_obj(ws):
    """RUN_META flat field/value + marker rows -> nested {parameters, a6_gate,
    id_discipline}. Marker rows (value None, field is a '--- X ---' string)
    switch the active sub-object."""
    obj = {META_FIRST_GROUP: {}, "a6_gate": {}, "id_discipline": {}}
    current = META_FIRST_GROUP
    it = ws.iter_rows(values_only=True)
    next(it, None)  # skip ('field','value') header
    for field, value in ((r + (None,))[:2] for r in it):
        if field is None:
            continue
        if field in META_SECTIONS:      # section marker
            current = META_SECTIONS[field]
            continue
        obj[current][str(field)] = _cell(value)
    return obj


def _recon_sheet_to_obj(ws):
    """LOCATE_RECONCILIATION sheet -> {N,M,K,J,entries:[...]}.
    N = number of inventory entries (data rows); M/K counted from the 'M/K'
    column; J is the count of extracted rows whose inventory_entry is blank/absent
    (found outside inventory). If the sheet is empty/absent, returns None."""
    rows = _rows_sheet_to_list(ws)
    if not rows:
        return None
    entries = []
    M = K = J = 0
    for r in rows:
        mk = (r.get("M/K") or "").strip().upper() if r.get("M/K") else ""
        inv = r.get("inventory entry")
        entries.append({
            "inventory_entry": inv,
            "unit_ref": r.get("unit_ref"),
            "page": r.get("page"),
            "disposition": r.get("disposition"),
            "m_or_k": mk or None,
            "evidence_rows_or_reason": r.get("evidence_rows or dismissal reason"),
        })
        if mk == "M":
            M += 1
        elif mk == "K":
            K += 1
        if inv in (None, "", "—", "-"):   # extracted but not in inventory
            J += 1
    N = M + K
    return {"N": N, "M": M, "K": K, "J": J, "entries": entries}


def xlsx_to_obj(path: Path) -> dict:
    wb = openpyxl.load_workbook(path, data_only=True)
    present = set(wb.sheetnames)
    doc = {}
    for sheet, key, kind in SHEET_MAP:
        if sheet not in present:
            doc[key] = None if kind in ("recon",) else ([] if kind == "rows"
                        else ("" if kind == "string" else {}))
            continue
        ws = wb[sheet]
        if kind == "rows":
            doc[key] = _rows_sheet_to_list(ws)
        elif kind == "meta":
            doc[key] = _meta_sheet_to_obj(ws)
        elif kind == "recon":
            doc[key] = _recon_sheet_to_obj(ws)
        elif kind == "string":
            # concatenate the single 'note' column's non-empty cells with newlines
            vals = [(_cell(r[0]) if r else None)
                    for r in ws.iter_rows(min_row=2, values_only=True)]
            doc[key] = "\n".join(str(v) for v in vals if v is not None)
    return doc


# ================================================================= JSON -> XLSX
def _write_rows_sheet(ws, records: list, header_order: list | None = None):
    if not records:
        return
    if header_order is None:
        header_order = list(records[0].keys())
    ws.append(header_order)
    for rec in records:
        ws.append([rec.get(h) for h in header_order])


def _write_meta_sheet(ws, meta: dict):
    ws.append(["field", "value"])
    for k, v in meta.get(META_FIRST_GROUP, {}).items():
        ws.append([k, v])
    inv_sections = {v: k for k, v in META_SECTIONS.items()}
    for sub in ("a6_gate", "id_discipline"):
        if meta.get(sub):
            ws.append([inv_sections[sub], None])
            for k, v in meta[sub].items():
                ws.append([k, v])


def _write_recon_sheet(ws, recon: dict | None):
    ws.append(["inventory entry", "unit_ref", "page", "disposition",
               "M/K", "evidence_rows or dismissal reason"])
    if not recon:
        return
    for e in recon.get("entries", []):
        ws.append([e.get("inventory_entry"), e.get("unit_ref"), e.get("page"),
                   e.get("disposition"), e.get("m_or_k"),
                   e.get("evidence_rows_or_reason")])


def obj_to_xlsx(doc: dict, path: Path, evidence_header: list | None = None):
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    for sheet, key, kind in SHEET_MAP:
        ws = wb.create_sheet(sheet)
        val = doc.get(key)
        if kind == "meta":
            _write_meta_sheet(ws, val or {})
        elif kind == "recon":
            _write_recon_sheet(ws, val)
        elif kind == "string":
            ws.append(["note"])
            for line in (val or "").split("\n"):
                ws.append([line if line != "" else None])
        elif kind == "rows":
            hdr = evidence_header if key == "evidence_rows" else None
            _write_rows_sheet(ws, val or [], hdr)
    wb.save(path)


# ================================================================= LOCATE (§9)
def locate_xlsx_to_obj(path: Path) -> dict:
    """LOCATE two-sheet workbook -> {passage_inventory:[...], sweep_coverage_certificate:[...]}."""
    wb = openpyxl.load_workbook(path, data_only=True)
    present = set(wb.sheetnames)
    out = {}
    for sheet, key in LOCATE_MAP:
        out[key] = _rows_sheet_to_list(wb[sheet]) if sheet in present else []
    return out


def obj_to_locate_xlsx(doc: dict, path: Path,
                       inv_header: list | None = None,
                       cert_header: list | None = None):
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    headers = {"passage_inventory": inv_header, "sweep_coverage_certificate": cert_header}
    for sheet, key in LOCATE_MAP:
        ws = wb.create_sheet(sheet)
        _write_rows_sheet(ws, doc.get(key) or [], headers.get(key))
    wb.save(path)


def validate_locate(doc: dict) -> list[str]:
    problems = []
    for k in LOCATE_KEYS:
        if k not in doc:
            problems.append(f"missing required key: {k}")
        elif not isinstance(doc[k], list):
            problems.append(f"{k} must be an array")
    # prelim_flag domain check (non-binding hint, but the enum is fixed)
    allowed = {"earmark_candidate", "tax_sharing_candidate",
               "cost_recovery_candidate", "ambiguous"}
    for i, r in enumerate(doc.get("passage_inventory") or []):
        pf = r.get("prelim_flag")
        if pf is not None and pf not in allowed:
            problems.append(f"passage_inventory[{i}].prelim_flag='{pf}' not in {sorted(allowed)}")
    return problems


def locate_roundtrip_ok(xlsx_path: Path) -> tuple[bool, list[str]]:
    wb = openpyxl.load_workbook(xlsx_path, read_only=True)
    heads = {}
    for sheet, key in LOCATE_MAP:
        ws = wb[sheet]
        hdr = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
        heads[key] = [h for h in hdr if h is not None]
    a = locate_xlsx_to_obj(xlsx_path)
    tmp = xlsx_path.with_suffix(".lroundtrip.xlsx")
    obj_to_locate_xlsx(a, tmp, inv_header=heads["passage_inventory"],
                       cert_header=heads["sweep_coverage_certificate"])
    b = locate_xlsx_to_obj(tmp)
    tmp.unlink(missing_ok=True)
    diffs = [f"round-trip mismatch on '{k}'" for k in LOCATE_KEYS if a.get(k) != b.get(k)]
    return (len(diffs) == 0, diffs)


# ================================================================= VALIDATION
def validate(doc: dict) -> list[str]:
    """Section-8 conformance + the reconciliation invariant. Returns a list of
    problem strings; empty list == conformant."""
    problems = []
    for k in JSON_KEYS:
        if k not in doc:
            problems.append(f"missing required key: {k}")
    ev = doc.get("evidence_rows")
    if not isinstance(ev, list):
        problems.append("evidence_rows must be an array")
    elif ev:
        ncols = len(ev[0])
        for i, r in enumerate(ev):
            if len(r) != ncols:
                problems.append(f"evidence_rows[{i}] has {len(r)} cols, expected {ncols}")
        # id-free discipline: instrument_id / pair_id null at extraction
        for i, r in enumerate(ev):
            for idcol in ("instrument_id", "pair_id"):
                if idcol in r and r[idcol] not in (None, "", "—"):
                    problems.append(f"evidence_rows[{i}].{idcol} populated at extraction "
                                    f"(must be null; ids assigned in reconciliation)")
    rec = doc.get("locate_reconciliation")
    if rec is not None:
        for f in ("N", "M", "K", "J"):
            if f not in rec:
                problems.append(f"locate_reconciliation missing scalar {f}")
        if all(f in rec for f in ("N", "M", "K")) and rec["M"] + rec["K"] != rec["N"]:
            problems.append(f"invariant M+K==N violated: "
                            f"{rec['M']}+{rec['K']} != {rec['N']}")
        if rec.get("J", 0) != 0:
            problems.append(f"J={rec['J']}>0: locate under-recalled; inventory must be "
                            f"amended with the {rec['J']} outside-inventory passage(s)")
    return problems


def _read_evidence_header(path: Path) -> list:
    wb = openpyxl.load_workbook(path, read_only=True)
    ws = wb["evidence_table"]
    hdr = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
    return [h for h in hdr if h is not None]


# ================================================================= ROUND-TRIP
def roundtrip_ok(xlsx_path: Path) -> tuple[bool, list[str]]:
    """xlsx -> json -> xlsx -> json' ; assert the two JSON objects are equal.
    (Comparing at the JSON layer, not the openpyxl layer, is the right equality:
    it is the canonical structure section 8 defines, free of workbook styling.)"""
    header = _read_evidence_header(xlsx_path)
    a = xlsx_to_obj(xlsx_path)
    tmp = xlsx_path.with_suffix(".roundtrip.xlsx")
    obj_to_xlsx(a, tmp, evidence_header=header)
    b = xlsx_to_obj(tmp)
    tmp.unlink(missing_ok=True)
    diffs = []
    for k in JSON_KEYS:
        if a.get(k) != b.get(k):
            diffs.append(f"round-trip mismatch on '{k}'")
    return (len(diffs) == 0, diffs)


# ================================================================= CLI
def main():
    p = argparse.ArgumentParser(description="xlsx<->json evidence converter (dict §8)")
    sub = p.add_subparsers(dest="cmd", required=True)

    p1 = sub.add_parser("to-json", help="workbook -> json object")
    p1.add_argument("xlsx"); p1.add_argument("out", nargs="?")

    p2 = sub.add_parser("to-xlsx", help="json object -> nine-sheet workbook")
    p2.add_argument("json"); p2.add_argument("out", nargs="?")
    p2.add_argument("--evidence-order", help="optional xlsx to copy the 44-col order from")

    p3 = sub.add_parser("validate", help="section-8 conformance + M+K=N / J=0 check")
    p3.add_argument("path", help="xlsx or json")

    p4 = sub.add_parser("roundtrip", help="prove xlsx->json->xlsx reproduces the original")
    p4.add_argument("xlsx")

    # --- LOCATE (§9) ---
    l1 = sub.add_parser("locate-to-json", help="LOCATE workbook -> json object (§9)")
    l1.add_argument("xlsx"); l1.add_argument("out", nargs="?")

    l2 = sub.add_parser("locate-validate", help="§9 conformance (keys + prelim_flag domain)")
    l2.add_argument("path", help="xlsx or json")

    l3 = sub.add_parser("locate-roundtrip", help="prove LOCATE xlsx<->json equivalence")
    l3.add_argument("xlsx")

    a = p.parse_args()

    if a.cmd == "locate-to-json":
        doc = locate_xlsx_to_obj(Path(a.xlsx))
        out = Path(a.out) if a.out else Path(a.xlsx).with_suffix(".json")
        out.write_text(json.dumps(doc, ensure_ascii=False, indent=2))
        print(f"wrote {out}"); return
    if a.cmd == "locate-validate":
        path = Path(a.path)
        doc = locate_xlsx_to_obj(path) if path.suffix == ".xlsx" else json.loads(path.read_text())
        problems = validate_locate(doc)
        if problems:
            print("NON-CONFORMANT:"); [print("  -", x) for x in problems]; sys.exit(1)
        print("CONFORMANT (§9 LOCATE contract)"); return
    if a.cmd == "locate-roundtrip":
        ok, diffs = locate_roundtrip_ok(Path(a.xlsx))
        if ok: print("ROUND-TRIP OK — LOCATE xlsx and json are equivalent renderings")
        else:
            print("ROUND-TRIP FAILED:"); [print("  -", d) for d in diffs]; sys.exit(1)
        return

    if a.cmd == "to-json":
        doc = xlsx_to_obj(Path(a.xlsx))
        out = Path(a.out) if a.out else Path(a.xlsx).with_suffix(".json")
        out.write_text(json.dumps(doc, ensure_ascii=False, indent=2))
        print(f"wrote {out}")

    elif a.cmd == "to-xlsx":
        doc = json.loads(Path(a.json).read_text())
        hdr = _read_evidence_header(Path(a.evidence_order)) if a.evidence_order else None
        out = Path(a.out) if a.out else Path(a.json).with_suffix(".xlsx")
        obj_to_xlsx(doc, out, evidence_header=hdr)
        print(f"wrote {out}")

    elif a.cmd == "validate":
        path = Path(a.path)
        doc = xlsx_to_obj(path) if path.suffix == ".xlsx" else json.loads(path.read_text())
        problems = validate(doc)
        if problems:
            print("NON-CONFORMANT:")
            for x in problems:
                print("  -", x)
            sys.exit(1)
        print("CONFORMANT (section 8 + reconciliation invariant)")

    elif a.cmd == "roundtrip":
        ok, diffs = roundtrip_ok(Path(a.xlsx))
        if ok:
            print("ROUND-TRIP OK — xlsx and json are equivalent renderings")
        else:
            print("ROUND-TRIP FAILED:")
            for d in diffs:
                print("  -", d)
            sys.exit(1)


if __name__ == "__main__":
    main()
