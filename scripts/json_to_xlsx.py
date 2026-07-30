#!/usr/bin/env python3
"""
json_to_xlsx.py — convert extraction JSON output to an Excel workbook.

Usage:
    python3 json_to_xlsx.py input.json [output.xlsx]
    python3 json_to_xlsx.py *.json            # batch: each foo.json -> foo.xlsx

Handles two JSON shapes:
  1. Nine-key EXTRACT object (dictionary section 8): run_meta + evidence_rows +
     the seven audit keys  -> nine-sheet workbook.
  2. A bare flat list of evidence rows           -> single 'evidence_table' sheet.

Each audit key becomes its own sheet. Scalars/strings (schema_stress_note) go
in a one-column sheet; nested objects (run_meta, locate_reconciliation) are
flattened to field/value rows so nothing is lost.
"""

import sys, json
from pathlib import Path
import openpyxl

# canonical sheet order for the nine-key object (dictionary section 8)
NINE_KEY_ORDER = [
    "run_meta", "evidence_rows", "coverage_report", "recall_audit",
    "locate_reconciliation", "gold_scoring", "field_exercise_note",
    "confidence_reasons", "schema_stress_note",
]
# json key -> worksheet name
SHEET_NAMES = {
    "run_meta": "RUN_META",
    "evidence_rows": "evidence_table",
    "coverage_report": "COVERAGE_REPORT",
    "recall_audit": "RECALL_AUDIT",
    "locate_reconciliation": "LOCATE_RECONCILIATION",
    "gold_scoring": "GOLD_SCORING",
    "field_exercise_note": "FIELD_EXERCISE_NOTE",
    "confidence_reasons": "CONFIDENCE_REASONS",
    "schema_stress_note": "SCHEMA_STRESS_NOTE",
}


def _flatten(obj, prefix=""):
    """Flatten a nested dict into [(dotted_key, value), ...] for field/value sheets."""
    rows = []
    if isinstance(obj, dict):
        for k, v in obj.items():
            key = f"{prefix}{k}"
            if isinstance(v, (dict, list)):
                rows += _flatten(v, prefix=f"{key}.")
            else:
                rows.append((key, v))
    elif isinstance(obj, list):
        for i, v in enumerate(obj):
            key = f"{prefix}{i}"
            if isinstance(v, (dict, list)):
                rows += _flatten(v, prefix=f"{key}.")
            else:
                rows.append((key, v))
    else:
        rows.append((prefix.rstrip("."), obj))
    return rows


def _cell_safe(v):
    """Excel cells cannot hold lists/dicts; serialize those to a compact string.
    Everything else passes through unchanged."""
    if isinstance(v, (list, dict)):
        return json.dumps(v, ensure_ascii=False)
    return v


def _write_rows_of_dicts(ws, records):
    """A list of dicts -> header row (union of keys, first-seen order) + one row each."""
    if not records:
        return
    header, seen = [], set()
    for r in records:
        for k in r.keys():
            if k not in seen:
                seen.add(k); header.append(k)
    ws.append(header)
    for r in records:
        ws.append([_cell_safe(r.get(h)) for h in header])


def _write_sheet(wb, key, value):
    name = SHEET_NAMES.get(key, key[:31])
    ws = wb.create_sheet(name[:31])
    # list of dicts -> table
    if isinstance(value, list) and value and isinstance(value[0], dict):
        _write_rows_of_dicts(ws, value)
    # list of scalars -> single column
    elif isinstance(value, list):
        ws.append([key])
        for v in value:
            ws.append([_cell_safe(v)])
    # nested object -> flattened field/value
    elif isinstance(value, dict):
        ws.append(["field", "value"])
        for k, v in _flatten(value):
            ws.append([k, _cell_safe(v)])
    # bare scalar/string -> one cell under a header
    else:
        ws.append([key])
        for line in str(value if value is not None else "").split("\n"):
            ws.append([line])


def convert(in_path: Path, out_path: Path):
    # Explicit UTF-8: on Windows, read_text() defaults to the system codepage
    # (e.g. cp1252), which turns clean UTF-8 accents into mojibake (é -> Ã©).
    doc = json.loads(in_path.read_text(encoding="utf-8"))
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    if isinstance(doc, list):
        # bare list of evidence rows
        ws = wb.create_sheet("evidence_table")
        _write_rows_of_dicts(ws, doc)
    elif isinstance(doc, dict):
        # nine-key object (in canonical order first, then any extras)
        keys = [k for k in NINE_KEY_ORDER if k in doc] + \
               [k for k in doc if k not in NINE_KEY_ORDER]
        for k in keys:
            _write_sheet(wb, k, doc[k])
    else:
        raise ValueError("Unsupported JSON top-level type")

    if not wb.sheetnames:                      # safety: never save an empty book
        wb.create_sheet("evidence_table")
    wb.save(out_path)
    return out_path


def _expand(args):
    """Expand any glob patterns ourselves so behaviour is identical on Windows
    (cmd/PowerShell do NOT expand *.json) and Unix shells (which do)."""
    import glob
    out = []
    for a in args:
        if any(ch in a for ch in "*?["):
            matches = glob.glob(a)
            if not matches:
                print(f"no files match: {a}")
            out.extend(matches)
        else:
            out.append(a)
    return out


def main():
    args = _expand(sys.argv[1:])
    if not args:
        print(__doc__); sys.exit(1)
    # single file with explicit output
    if len(args) == 2 and args[0].endswith(".json") and args[1].endswith(".xlsx"):
        out = convert(Path(args[0]), Path(args[1]))
        print(f"wrote {out}")
        return
    # batch: one or more json files, each -> same-stem .xlsx
    for a in args:
        p = Path(a)
        if p.suffix != ".json":
            print(f"skip (not .json): {a}"); continue
        out = p.with_suffix(".xlsx")
        convert(p, out)
        print(f"wrote {out}")


if __name__ == "__main__":
    main()
